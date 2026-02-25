from flask import flash
from app.models import Subject, Faculty
from app import db
from datetime import datetime
import io
import re


def generate_sample_file():
    """Talabalarni import qilish uchun namuna Excel fayl (yangi tartib)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar import"

    # Sarlavha
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = "Talabalar import uchun namuna fayl"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Import talablari
    from datetime import datetime
    ws['A2'] = "IMPORT TALABLARI:"
    ws.merge_cells('A2:P2')
    ws['A2'].font = Font(size=11, bold=True, color="000000")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    
    # Talablar ro'yxati
    requirements = [
        "1. Talaba ID - majburiy maydon, unikal bo'lishi kerak",
        "2. To'liq ism - majburiy maydon",
        "3. Pasport seriya raqami - majburiy maydon (masalan: AB1234567)",
        "4. JSHSHIR - ixtiyoriy maydon (14 raqam)",
        "5. Tug'ilgan sana - ixtiyoriy maydon (DD.MM.YYYY yoki YYYY-MM-DD formatida)",
        "6. Telefon - ixtiyoriy maydon",
        "7. Email - ixtiyoriy maydon, unikal bo'lishi kerak",
        "8. Tavsif - ixtiyoriy maydon",
        "9. Fakultet - ixtiyoriy maydon (guruh biriktirish uchun)",
        "10. Kurs - ixtiyoriy maydon (1-kurs, 2-kurs formatida, guruh biriktirish uchun)",
        "11. Semestr - ixtiyoriy maydon (1-semestr, 2-semestr formatida)",
        "12. Ta'lim shakli - ixtiyoriy maydon (Kunduzgi, Sirtqi, Kechki, Masofaviy - bosh harf katta, guruh biriktirish uchun)",
        "13. Qabul yili - ixtiyoriy maydon (masalan: 2024, guruh biriktirish uchun)",
        "14. Mutaxassislik kodi - ixtiyoriy maydon (yo'nalish kodi)",
        "15. Mutaxassislik nomi - ixtiyoriy maydon (yo'nalish nomi)",
        "16. Guruh - ixtiyoriy maydon (agar mavjud bo'lsa, qo'shiladi, aks holda yangi yaratiladi)"
    ]
    
    for idx, req in enumerate(requirements, start=3):
        ws.merge_cells(f'A{idx}:P{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = req
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

    # Eslatma qismi
    note_start_row = len(requirements) + 3
    ws.merge_cells(f'A{note_start_row}:O{note_start_row}')
    note_title_cell = ws.cell(row=note_start_row, column=1)
    note_title_cell.value = "ESLATMA:"
    note_title_cell.font = Font(size=11, bold=True, color="000000")
    note_title_cell.alignment = Alignment(horizontal='left', vertical='center')
    note_title_cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    
    notes = [
        "• Fayl .xlsx yoki .xls formatida bo'lishi kerak",
        "• Majburiy maydonlar: To'liq ism, Talaba ID, Pasport seriya raqami",
        "• Ixtiyoriy maydonlar: Qolgan barcha maydonlar",
        "• Agar Fakultet, Kurs, Ta'lim shakli, Qabul yili va Guruh ma'lumotlari to'liq bo'lsa, talaba avtomatik guruhga qo'shiladi",
        "• Guruh nomi to'g'ri yozilishi kerak (masalan: DI-21). Agar guruh mavjud bo'lmasa, yangi yaratiladi",
        "• Qabul yili yo'nalishga biriktirishda muhim (masalan: 2024, 2025)",
        "• Email va Talaba ID takrorlanmasligi kerak",
        "• Yangi talabalar uchun boshlang'ich parol: Pasport seriya raqami"
    ]
    
    for idx, note in enumerate(notes, start=note_start_row + 1):
        ws.merge_cells(f'A{idx}:P{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = note
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = [
        "Talaba ID",              # A
        "To'liq ism",             # B
        "Pasport seriya raqami",  # C
        "JSHSHIR",                # D
        "Tug'ilgan sana",         # E
        "Telefon",                # F
        "Email",                  # G
        "Tavsif",                 # H
        "Fakultet",               # I
        "Kurs",                   # J
        "Semestr",                # K
        "Ta'lim shakli",          # L
        "Qabul yili",             # M
        "Mutaxassislik kodi",     # N
        "Mutaxassislik nomi",     # O
        "Guruh"                   # P
    ]

    header_row = len(requirements) + len(notes) + 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Namuna ma'lumotlar (ismlar katta harflarda)
    sample_data = [
        ["ST2024001", "ALIYEV VALI", "AB1234567", "30202020200021", "2000-01-15", "+998901234567", "vali@example.com", "Talaba haqida ma'lumot", "IT", "1-kurs", "1-semestr", "Kunduzgi", "2024", "DI", "Dasturiy injiniring", "DI-21"],
        ["ST2024002", "KARIMOVA ZUHRA", "AC2345678", "30202020200022", "2001-03-20", "+998901234568", "zuhra@example.com", "Talaba haqida ma'lumot", "IT", "1-kurs", "1-semestr", "Kunduzgi", "2024", "DI", "Dasturiy injiniring", "DI-21"]
    ]

    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Ustun kengliklarini sozlash
    column_widths = [15, 30, 20, 18, 18, 16, 25, 40, 20, 12, 12, 15, 12, 20, 30, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_staff_sample_file():
    """Xodimlarni import qilish uchun namuna Excel fayl (bitta sheet'da)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Xodimlar"
    
    # Sarlavha
    title = "Xodimlar import uchun namuna fayl"
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Import talablari
    from datetime import datetime
    ws['A2'] = "IMPORT TALABLARI:"
    ws.merge_cells('A2:I2')
    ws['A2'].font = Font(size=11, bold=True, color="000000")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    
    # Talablar ro'yxati
    requirements = [
        "1. To'liq ism - majburiy maydon",
        "2. Login - majburiy maydon, unikal bo'lishi kerak",
        "3. Pasport seriya raqami - majburiy maydon (masalan: AB1234567)",
        "4. JSHSHIR - ixtiyoriy maydon (14 raqam)",
        "5. Tug'ilgan sana - ixtiyoriy maydon (DD.MM.YYYY yoki YYYY-MM-DD formatida)",
        "6. Telefon - ixtiyoriy maydon",
        "7. Email - ixtiyoriy maydon, unikal bo'lishi kerak",
        "8. Tavsif - ixtiyoriy maydon. Dekan roli bo'lsa, fakultet nomi yozilishi kerak (masalan: IT fakulteti dekani)",
        "9. Rollar - majburiy maydon (kamida bitta rol tanlanishi kerak). Vergul bilan ajratilgan: Administrator, Dekan, O'qituvchi, Buxgalter"
    ]
    
    for idx, req in enumerate(requirements, start=3):
        ws.merge_cells(f'A{idx}:I{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = req
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

    # Eslatma qismi
    note_start_row = len(requirements) + 3
    ws.merge_cells(f'A{note_start_row}:I{note_start_row}')
    note_title_cell = ws.cell(row=note_start_row, column=1)
    note_title_cell.value = "ESLATMA:"
    note_title_cell.font = Font(size=11, bold=True, color="000000")
    note_title_cell.alignment = Alignment(horizontal='left', vertical='center')
    note_title_cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    
    notes = [
        "• Fayl .xlsx yoki .xls formatida bo'lishi kerak",
        "• Majburiy maydonlar: To'liq ism, Login, Pasport seriya raqami, Rollar (kamida bitta)",
        "• Ixtiyoriy maydonlar: Qolgan barcha maydonlar",
        "• Dekan roli tanlangan bo'lsa, Tavsif maydonida fakultet nomi ko'rsatilishi kerak (masalan: IT fakulteti dekani)",
        "• Login va Email takrorlanmasligi kerak",
        "• Yangi xodimlar uchun boshlang'ich parol: Pasport seriya raqami"
    ]
    
    for idx, note in enumerate(notes, start=note_start_row + 1):
        ws.merge_cells(f'A{idx}:I{idx}')
        cell = ws.cell(row=idx, column=1)
        cell.value = note
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    
    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = ["To'liq ism", 'Login', 'Pasport seriya raqami', 'JSHSHIR', "Tug'ilgan sana", 'Telefon', 'Email', 'Tavsif', 'Rollar']
    header_row = len(requirements) + len(notes) + 4
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Namuna ma'lumotlar (ismlar katta harflarda)
    sample_data = [
        ["TURSUNQULOV AVAZBEK", "admin", "AB1234567", "30202020200021", "1980-01-15", "+998901234567", "admin@university.uz", "Tizim administratori", "Administrator"],
        ["KARIMOV SHERZOD", "sherzod", "AC2345678", "30202020200022", "1975-03-20", "+998901234568", "dean.it@university.uz", "IT fakulteti dekani", "Dekan"],
        ["MAMATOV VALIJON", "valijon", "AD3456789", "30202020200023", "1985-05-10", "+998901234569", "valijon@university.uz", "Dasturiy injiniring kafedrasi o'qituvchisi", "O'qituvchi"],
        ["RAHIMOVA AZIZA", "aziza", "AE4567890", "30202020200024", "1990-07-25", "+998901234570", "accounting@university.uz", "Buxgalteriya bo'limi xodimi", "Buxgalter"],
        ["ALIYEV VALI", "vali", "AF5678901", "30202020200025", "1982-09-30", "+998901234571", "vali@university.uz", "IT fakulteti dekani va o'qituvchi", "Dekan, O'qituvchi"]
    ]
    
    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Ustun kengliklarini sozlash
    column_widths = [30, 20, 20, 18, 18, 16, 25, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_students_from_excel(file, faculty_id=None):
    """Excel fayldan talabalarni import qilish (yangi tartib)
    
    Args:
        file: Excel fayl
        faculty_id: Fakultet ID (ixtiyoriy, agar berilsa, guruhlar shu fakultet doirasida qidiriladi)
    """
    try:
        from openpyxl import load_workbook
        from app.models import User, Group, Faculty, Direction
        from app import db
        from datetime import datetime, date
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        updated = 0
        errors = []
        
        # Sarlavha qatorini topish (dinamik ravishda)
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            first_cell = ws.cell(row=row_num, column=1).value
            if first_cell and ("Talaba ID" in str(first_cell) or "To'liq ism" in str(first_cell)):
                header_row = row_num
                break
        
        if not header_row:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': ["Sarlavha qatori topilmadi. Iltimos, fayl formati to'g'ri ekanligini tekshiring."]
            }
        
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get("To'liq ism") and not row_data.get('Email'):
                    continue
                
                full_name = row_data.get("To'liq ism", '').strip()
                student_id = row_data.get('Talaba ID', '').strip()
                email = row_data.get('Email', '').strip()
                passport_number = row_data.get('Pasport seriya raqami', '').strip()
                
                # Majburiy maydonlarni tekshirish
                if not full_name:
                    errors.append(f"Qator {row_num}: To'liq ism kiritilmagan")
                    continue
                if not student_id:
                    errors.append(f"Qator {row_num}: Talaba ID kiritilmagan")
                    continue
                if not passport_number:
                    errors.append(f"Qator {row_num}: Pasport seriya raqami kiritilmagan")
                    continue
                
                # Ismni katta harflarda yozish
                full_name = full_name.upper()
                
                # Foydalanuvchini topish (student_id, email yoki passport_number orqali)
                user = None
                if student_id:
                    user = User.query.filter_by(student_id=student_id).first()
                if not user and email:
                    user = User.query.filter_by(email=email).first()
                if not user and passport_number:
                    user = User.query.filter_by(passport_number=passport_number).first()
                
                # JSHSHIR
                pinfl = row_data.get('JSHSHIR', '').strip() or None
                
                # Tug'ilgan sana
                birth_date_str = row_data.get("Tug'ilgan sana", '').strip()
                birth_date = None
                if birth_date_str:
                    try:
                        # DD.MM.YYYY yoki YYYY-MM-DD formatini qo'llab-quvvatlash
                        if '.' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y').date()
                        elif '-' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format (DD.MM.YYYY yoki YYYY-MM-DD)")
                
                # Fakultet, Kurs, Semestr, Ta'lim shakli, Qabul yili, Mutaxassislik, Guruh
                faculty_name = row_data.get('Fakultet', '').strip()
                course_str = row_data.get('Kurs', '').strip()  # "1-kurs" formatida
                semester_str = row_data.get('Semestr', '').strip()
                education_type = row_data.get("Ta'lim shakli", '').strip()
                enrollment_year_str = row_data.get('Qabul yili', '').strip()
                specialty_code = row_data.get('Mutaxassislik kodi', '').strip()
                specialty_name = row_data.get('Mutaxassislik nomi', '').strip()
                group_name = row_data.get('Guruh', '').strip()
                
                # Kurs raqamini ajratish ("1-kurs" -> 1)
                course_year = None
                if course_str:
                    try:
                        course_year = int(course_str.replace('-kurs', '').strip())
                    except:
                        pass
                
                # Semestr raqamini ajratish ("1-semestr" -> 1)
                semester = None
                if semester_str:
                    try:
                        # "1-semestr" formatidan raqamni ajratish
                        semester_str_clean = semester_str.replace('-semestr', '').strip()
                        semester = int(semester_str_clean)
                    except:
                        pass
                
                # Ta'lim shaklini kichik harfga o'tkazish (database'da kichik harfda saqlanadi)
                if education_type:
                    education_type = education_type.lower()
                
                # Qabul yilini o'qish
                enrollment_year = None
                if enrollment_year_str:
                    try:
                        enrollment_year = int(enrollment_year_str)
                    except:
                        pass
                
                # Yo'nalishni topish yoki yaratish
                direction = None
                if specialty_code and faculty_name:
                    # Fakultetni topish
                    faculty = Faculty.query.filter_by(name=faculty_name).first()
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                        continue
                    
                    # Yo'nalishni topish (kodi bo'yicha va qabul yili bo'yicha)
                    query = Direction.query.filter_by(code=specialty_code, faculty_id=faculty.id)
                    if enrollment_year:
                        # Qabul yili bilan bir nechta yo'nalish bo'lishi mumkin
                        query = query.filter_by(enrollment_year=enrollment_year)
                    direction = query.first()
                    
                    # Agar qabul yili bilan topilmasa, qabul yilisiz qidirish
                    if not direction and enrollment_year:
                        direction = Direction.query.filter_by(
                            code=specialty_code, 
                            faculty_id=faculty.id,
                            enrollment_year=None
                        ).first()
                    
                    # Agar yo'nalish topilmasa, yangi yaratish
                    if not direction:
                        if specialty_name:
                            direction = Direction(
                                name=specialty_name,
                                code=specialty_code,
                                faculty_id=faculty.id
                            )
                            db.session.add(direction)
                            db.session.flush()
                    pass
                
                # Guruhni topish yoki yaratish
                group = None
                if group_name and faculty_name:
                    # Fakultetni topish (agar yo'nalishda topilmagan bo'lsa)
                    if not direction:
                        faculty = Faculty.query.filter_by(name=faculty_name).first()
                        if not faculty:
                            errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                            continue
                    else:
                        faculty = direction.faculty
                    
                    # Guruhni topish
                    group = Group.query.filter_by(name=group_name, faculty_id=faculty.id).first()
                    
                    # Agar guruh topilmasa, yangi yaratish
                    if not group:
                        if course_year and education_type:
                            group = Group(
                                name=group_name,
                                faculty_id=faculty.id,
                                direction_id=direction.id if direction else None,
                                course_year=course_year,
                                semester=semester or 1,
                                education_type=education_type,
                                enrollment_year=enrollment_year
                            )
                            db.session.add(group)
                            db.session.flush()
                        else:
                            errors.append(f"Qator {row_num}: Guruh yaratish uchun kurs va ta'lim shakli kerak")
                    elif direction and not group.direction_id:
                        # Agar guruh topilgan bo'lsa va yo'nalish biriktirilmagan bo'lsa, biriktirish
                        group.direction_id = direction.id
                
                if user:
                    # Yangilash
                    user.full_name = full_name
                    if student_id:
                        user.student_id = student_id
                    user.phone = row_data.get('Telefon', '').strip() or None
                    user.passport_number = passport_number
                    user.pinfl = pinfl
                    user.birth_date = birth_date
                    user.email = email if email else None
                    user.description = row_data.get('Tavsif', '').strip() or None
                    
                    # Guruhni biriktirish
                    if group:
                        user.group_id = group.id
                        if semester:
                            user.semester = semester
                        if education_type:
                            user.education_type = education_type
                        if enrollment_year:
                            user.enrollment_year = enrollment_year
                    
                    user.set_password(passport_number)
                    updated += 1
                else:
                    # Yaratish
                    user = User(
                        full_name=full_name,
                        role='student',
                        student_id=student_id or None,
                        phone=row_data.get('Telefon', '').strip() or None,
                        passport_number=passport_number,
                        pinfl=pinfl,
                        birth_date=birth_date,
                        email=email if email else None,
                        description=row_data.get('Tavsif', '').strip() or None,
                        semester=semester,
                        education_type=education_type if education_type else None,
                        enrollment_year=enrollment_year
                    )
                    
                    # Guruhni biriktirish
                    if group:
                        user.group_id = group.id
                    
                    user.set_password(passport_number)
                    db.session.add(user)
                    
                    # Commit qilish va agar email NOT NULL xatolik bo'lsa, email maydonini bo'sh qatorga o'zgartirish
                    try:
                        db.session.flush()  # ID olish uchun
                    except Exception as e:
                        error_str = str(e).lower()
                        if 'email' in error_str and ('not null' in error_str or 'constraint' in error_str):
                            # Database'da email NOT NULL bo'lsa, bo'sh qator qo'yamiz
                            db.session.rollback()
                            user.email = ''  # Bo'sh qator (database NOT NULL constraint uchun)
                            db.session.add(user)
                            db.session.flush()
                        else:
                            raise
                    
                    imported += 1
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_directions_from_excel(file):
    """Excel fayldan yo'nalishlar va guruhlarni import qilish"""
    try:
        from openpyxl import load_workbook
        from app.models import Direction, Group, Faculty
        from app import db
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        imported = 0
        errors = []
        
        # Sarlavha qatorini topish (1-qator)
        header_row = 1
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get('Yo\'nalish nomi') and not row_data.get('Yo\'nalish kodi'):
                    continue
                
                direction_name = row_data.get('Yo\'nalish nomi', '').strip()
                direction_code = row_data.get('Yo\'nalish kodi', '').strip()
                faculty_name = row_data.get('Fakultet', '').strip()
                group_name = row_data.get('Guruh', '').strip()
                course_year = row_data.get('Kurs', '').strip()
                
                if not direction_name or not direction_code:
                    errors.append(f"Qator {row_num}: Yo'nalish nomi yoki kodi kiritilmagan")
                    continue
                
                # Fakultetni topish
                faculty = None
                if faculty_name:
                    faculty = Faculty.query.filter_by(name=faculty_name).first()
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet '{faculty_name}' topilmadi")
                        continue
                
                # Yo'nalishni topish yoki yaratish
                direction = Direction.query.filter_by(code=direction_code).first()
                if not direction:
                    if not faculty:
                        errors.append(f"Qator {row_num}: Fakultet kiritilmagan")
                        continue
                    direction = Direction(
                        name=direction_name,
                        code=direction_code,
                        description=row_data.get('Tavsif', '').strip() or None,
                        faculty_id=faculty.id
                    )
                    db.session.add(direction)
                    db.session.flush()
                    imported += 1
                
                # Guruhni topish yoki yaratish
                if group_name and faculty:
                    group = Group.query.filter_by(name=group_name).first()
                    if not group:
                        try:
                            course_year_int = int(course_year) if course_year else 1
                        except:
                            course_year_int = 1
                        
                        group = Group(
                            name=group_name,
                            faculty_id=faculty.id,
                            direction_id=direction.id,
                            course_year=course_year_int
                        )
                        db.session.add(group)
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_staff_from_excel(file):
    """Excel fayldan xodimlarni import qilish (bitta sheet'dan) - bir nechta rollarni qo'llab-quvvatlash"""
    try:
        from openpyxl import load_workbook
        from app.models import User, Faculty, UserRole
        from app import db
        from datetime import datetime, date
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        wb = load_workbook(file)
        ws = wb.active  # Bitta sheet'dan o'qish
        
        imported = 0
        updated = 0
        errors = []
        
        # Sarlavha qatorini topish (dinamik ravishda)
        header_row = None
        for row_num in range(1, min(20, ws.max_row + 1)):  # Birinchi 20 qatorni tekshirish
            first_cell = ws.cell(row=row_num, column=1).value
            if first_cell and ("To'liq ism" in str(first_cell) or "To'liq ismi" in str(first_cell)):
                header_row = row_num
                break
        
        if not header_row:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': ["Sarlavha qatori topilmadi. Iltimos, fayl formati to'g'ri ekanligini tekshiring."]
            }
        
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row_data.get("To'liq ism") and not row_data.get('Email'):
                    continue
                
                full_name = row_data.get("To'liq ism", '').strip()
                login = row_data.get('Login', '').strip()
                email = row_data.get('Email', '').strip()
                
                # Majburiy maydonlarni tekshirish
                if not full_name:
                    errors.append(f"Qator {row_num}: To'liq ism kiritilmagan")
                    continue
                if not login:
                    errors.append(f"Qator {row_num}: Login kiritilmagan")
                    continue
                
                # Xodim uchun
                passport_number = row_data.get('Pasport seriya raqami', '').strip()
                if not passport_number:
                    errors.append(f"Qator {row_num}: Pasport seriya raqami kiritilmagan")
                    continue
                
                # Ismni katta harflarda yozish
                full_name = full_name.upper()
                
                # Login yoki email orqali foydalanuvchini topish
                user = None
                if login:
                    user = User.query.filter_by(login=login).first()
                if not user and email:
                    user = User.query.filter_by(email=email).first()
                
                # Pasport raqamini katta harfga o'zgartirish
                passport_number = passport_number.upper()
                
                # JSHSHIR
                pinfl = row_data.get('JSHSHIR', '').strip() or None
                
                # Tug'ilgan sana
                birth_date_str = row_data.get("Tug'ilgan sana", '').strip()
                birth_date = None
                if birth_date_str:
                    try:
                        # YYYY-MM-DD formatini qo'llab-quvvatlash (asosiy)
                        # DD.MM.YYYY formatini ham qo'llab-quvvatlash (qayta ishlash uchun)
                        if '-' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                        elif '.' in birth_date_str:
                            birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y').date()
                        else:
                            errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format (YYYY-MM-DD, masalan: 1999-02-06)")
                    except ValueError:
                        errors.append(f"Qator {row_num}: Tug'ilgan sana noto'g'ri format (YYYY-MM-DD, masalan: 1999-02-06)")
                
                # Rollarni o'qish va qayta ishlash
                roles_str = row_data.get('Rollar', '').strip()
                roles_list = []
                if roles_str:
                    # O'zbek tilidagi rollarni ingliz tiliga o'girish
                    role_mapping = {
                        'administrator': 'admin',
                        'dekan': 'dean',
                        'o\'qituvchi': 'teacher',
                        'oqituvchi': 'teacher',
                        'buxgalter': 'accounting',
                        'talaba': 'student'
                    }
                    # Vergul bilan ajratilgan rollarni ajratish
                    roles_raw = [r.strip() for r in roles_str.split(',') if r.strip()]
                    # Har bir rolni tekshirish va o'girish
                    for role_raw in roles_raw:
                        role_lower = role_raw.lower()
                        # Avval to'g'ridan-to'g'ri ingliz tilida bo'lsa
                        if role_lower in ['admin', 'dean', 'teacher', 'accounting', 'student']:
                            roles_list.append(role_lower)
                        # Keyin o'zbek tilidagi rollarni o'girish
                        elif role_lower in role_mapping:
                            roles_list.append(role_mapping[role_lower])
                        # Agar hech qaysi biriga mos kelmasa, o'zbek tilidagi nomlarni tekshirish
                        else:
                            # Katta harflar bilan ham tekshirish
                            role_title = role_raw.title()
                            if role_title == 'Administrator':
                                roles_list.append('admin')
                            elif role_title == 'Dekan':
                                roles_list.append('dean')
                            elif role_title in ["O'qituvchi", 'Oqituvchi']:
                                roles_list.append('teacher')
                            elif role_title == 'Buxgalter':
                                roles_list.append('accounting')
                            elif role_title == 'Talaba':
                                roles_list.append('student')
                    
                    # Faqat to'g'ri rollarni qabul qilish
                    valid_roles = ['admin', 'dean', 'teacher', 'accounting', 'student']
                    roles_list = [r for r in roles_list if r in valid_roles]
                
                # Asosiy rol - agar rollar berilgan bo'lsa, eng yuqori darajali rolni tanlash
                # Aks holda default 'teacher' rolini beramiz
                if roles_list:
                    # Rol darajalari (yuqoridan pastga)
                    role_priority = ['admin', 'dean', 'teacher', 'accounting', 'student']
                    primary_role = None
                    for priority_role in role_priority:
                        if priority_role in roles_list:
                            primary_role = priority_role
                            break
                    if not primary_role:
                        primary_role = roles_list[0]
                else:
                    primary_role = 'teacher'
                    roles_list = ['teacher']  # Default rol
                
                # Dekan roli bo'lsa, fakultetni tavsifdan olish
                faculty_id_for_dean = None
                if 'dean' in roles_list:
                    description = row_data.get('Tavsif', '').strip()
                    if description:
                        # Tavsifdan fakultet nomini topish
                        faculty = Faculty.query.filter(Faculty.name.ilike(f'%{description}%')).first()
                        if not faculty:
                            # Aksincha, description fakultet nomi bo'lishi mumkin
                            faculty = Faculty.query.filter_by(name=description).first()
                        if faculty:
                            faculty_id_for_dean = faculty.id
                        else:
                            errors.append(f"Qator {row_num}: Dekan roli uchun tavsif maydonida fakultet nomi ko'rsatilishi kerak")
                            continue
                    else:
                        errors.append(f"Qator {row_num}: Dekan roli tanlangan bo'lsa, tavsif maydonida fakultet nomi ko'rsatilishi kerak")
                        continue
                
                if user:
                    # Yangilash
                    user.full_name = full_name
                    if login:
                        user.login = login
                    user.phone = row_data.get('Telefon', '').strip() or None
                    user.passport_number = passport_number
                    user.pinfl = pinfl
                    user.birth_date = birth_date
                    user.email = email if email else None
                    user.description = row_data.get('Tavsif', '').strip() or None
                    
                    # Dekan roli bo'lsa, fakultetni biriktirish
                    if 'dean' in roles_list and faculty_id_for_dean:
                        user.faculty_id = faculty_id_for_dean
                    elif 'dean' not in roles_list:
                        user.faculty_id = None
                    
                    user.set_password(passport_number)
                    
                    # Rollarni yangilash (commit qilmasdan)
                    user.role = primary_role
                    # Eski rollarni o'chirish
                    UserRole.query.filter_by(user_id=user.id).delete()
                    # Yangi rollarni qo'shish (belgilangan tartibda: admin, dean, teacher, accounting, student)
                    role_order = ['admin', 'dean', 'teacher', 'accounting', 'student']
                    if roles_list:
                        # Belgilangan tartibda qo'shish
                        for role in role_order:
                            if role in roles_list:
                                user_role = UserRole(user_id=user.id, role=role)
                                db.session.add(user_role)
                    else:
                        # Agar rollar berilmagan bo'lsa, faqat asosiy rolni saqlash
                        user_role = UserRole(user_id=user.id, role=primary_role)
                        db.session.add(user_role)
                    
                    updated += 1
                else:
                    # Yaratish
                    # Login unikalligi
                    if User.query.filter_by(login=login).first():
                        errors.append(f"Qator {row_num}: Bu login allaqachon mavjud")
                        continue
                    
                    # Email unikalligi (agar berilsa)
                    if email and User.query.filter_by(email=email).first():
                        errors.append(f"Qator {row_num}: Bu email allaqachon mavjud")
                        continue
                    
                    user = User(
                        login=login,
                        full_name=full_name,
                        role=primary_role,
                        phone=row_data.get('Telefon', '').strip() or None,
                        passport_number=passport_number,
                        pinfl=pinfl,
                        birth_date=birth_date,
                        description=row_data.get('Tavsif', '').strip() or None,
                        faculty_id=faculty_id_for_dean if 'dean' in roles_list else None
                    )
                    
                    # Email maydonini alohida o'rnatish (agar bo'sh bo'lsa, o'rnatmaymiz)
                    if email:
                        user.email = email
                    
                    # Parolni pasport raqamiga o'rnatish
                    user.set_password(passport_number)
                    
                    db.session.add(user)
                    
                    # Commit qilish va agar email NOT NULL xatolik bo'lsa, email maydonini bo'sh qatorga o'zgartirish
                    try:
                        db.session.flush()  # ID olish uchun
                    except Exception as e:
                        error_str = str(e).lower()
                        if 'email' in error_str and ('not null' in error_str or 'constraint' in error_str):
                            # Database'da email NOT NULL bo'lsa, bo'sh qator qo'yamiz
                            db.session.rollback()
                            user.email = ''  # Bo'sh qator (database NOT NULL constraint uchun)
                            db.session.add(user)
                            db.session.flush()
                        else:
                            raise
                    
                    # Rollarni biriktirish (user ID olish uchun flush kerak, lekin commit qilmaymiz)
                    # Rollarni belgilangan tartibda saqlash: admin, dean, teacher, accounting, student
                    role_order = ['admin', 'dean', 'teacher', 'accounting', 'student']
                    if roles_list:
                        # Belgilangan tartibda qo'shish
                        for role in role_order:
                            if role in roles_list:
                                user_role = UserRole(user_id=user.id, role=role)
                                db.session.add(user_role)
                    else:
                        user_role = UserRole(user_id=user.id, role=primary_role)
                        db.session.add(user_role)
                    
                    imported += 1
                
            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_all_users_from_excel(file):
    """Excel fayldan barcha foydalanuvchilarni import qilish (rol bo'yicha ajratish) - eski funksiya, import_staff_from_excel ishlatiladi"""
    return import_staff_from_excel(file)


def generate_subjects_sample_file():
    """Fanlarni import qilish uchun namuna Excel fayl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Fanlar import"

    # Sarlavha
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "Fanlar import uchun namuna fayl"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Jadval sarlavhalari
    headers = [
        "Fan nomi",      # A
        "Tavsif"         # B
    ]

    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Namuna ma'lumotlar
    sample_data = [
        ["Dasturlash asoslari", "Dasturlashning asosiy tushunchalari va algoritmlar"],
        ["Ma'lumotlar bazasi", "Ma'lumotlar bazasi dizayni va SQL so'rovlari"],
        ["Web dasturlash", "Web texnologiyalari va frameworklar"]
    ]

    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Ustun kengliklarini sozlash
    column_widths = [40, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_subjects_from_excel(file):
    """Excel fayldan fanlarni import qilish"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    imported = 0
    updated = 0
    errors = []

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Sarlavha qatorini topish (3-qator)
        header_row = 3
        headers = {}
        for col_num in range(1, 3):  # A, B ustunlari
            cell_value = ws.cell(row=header_row, column=col_num).value
            if cell_value:
                headers[cell_value] = col_num

        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                # Ustunlardan ma'lumotlarni olish
                name = ws.cell(row=row_num, column=headers.get("Fan nomi", 1)).value
                description = ws.cell(row=row_num, column=headers.get("Tavsif", 2)).value

                # Bo'sh qatorlarni o'tkazib yuborish
                if not name:
                    continue

                name = str(name).strip()
                description = str(description).strip() if description else None

                # Fan nomi bo'yicha tekshirish
                existing_subject = Subject.query.filter_by(name=name).first()

                if existing_subject:
                    # Yangilash
                    existing_subject.description = description
                    updated += 1
                else:
                    # Yaratish
                    subject = Subject(
                        name=name,
                        code='',  # NOT NULL constraint uchun
                        description=description,
                        credits=3,  # Default
                        semester=1  # Default
                    )
                    db.session.add(subject)
                    imported += 1

            except Exception as e:
                errors.append(f"Qator {row_num}: Xatolik - {str(e)}")

        db.session.commit()

        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors
        }

    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def import_curriculum_from_excel(file, direction_id, enrollment_year=None, education_type=None):
    """Excel fayldan o'quv rejani import qilish (rasmdagi formatga mos)"""
    try:
        from openpyxl import load_workbook
        from app.models import Direction, Subject, DirectionCurriculum
        from app import db
    except ImportError:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': ["openpyxl kutubxonasi o'rnatilmagan"]
        }
    
    try:
        direction = Direction.query.get(direction_id)
        if not direction:
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': ["Yo'nalish topilmadi"]
            }
        
        wb = load_workbook(file, data_only=True)
        
        imported = 0
        updated = 0
        subjects_created = 0
        errors = []
        
        # Birinchi worksheet'ni olish
        ws = wb.active
        
        # Sarlavha qatorini topish
        header_row = None
        for row_num in range(1, min(10, ws.max_row + 1)):
            first_cell = ws.cell(row=row_num, column=1).value
            if first_cell and ("Semestr" in str(first_cell) or "Fan nomi" in str(first_cell)):
                header_row = row_num
                break
        
        if not header_row:
            errors.append("Sarlavha qatori topilmadi.")
            return {
                'success': False,
                'imported': 0,
                'updated': 0,
                'errors': errors
            }
        
        # Sarlavhalarni olish
        headers = {}
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_num).value
            if cell_value:
                headers[str(cell_value).strip()] = col_num
        
        # Ma'lumotlarni o'qish
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                # Semestr (1-semestr formatida yoki faqat raqam)
                semester_cell = ws.cell(row=row_num, column=headers.get("Semestr", 1)).value
                if not semester_cell:
                    continue
                
                # Semestr raqamini olish
                semester_str = str(semester_cell).strip()
                if '-semestr' in semester_str:
                    semester = int(semester_str.split('-')[0])
                else:
                    semester = int(float(semester_str))
                
                # Fan nomi
                subject_name = ws.cell(row=row_num, column=headers.get("Fan nomi", 2)).value
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not subject_name:
                    continue
                
                # Fan topish yoki yaratish
                subject_name = str(subject_name).strip()
                subject = Subject.query.filter_by(name=subject_name).first()
                
                if not subject:
                    # Yangi fan yaratish
                    subject = Subject(
                        name=subject_name,
                        code='',
                        description=f"Import qilingan fan: {subject_name}",
                        credits=3,
                        semester=semester
                    )
                    db.session.add(subject)
                    db.session.flush()  # ID olish uchun
                    subjects_created += 1
                
                # Soatlar (yangi tartibda: Semestr, Fan nomi, Maruza, Amaliyot, ...)
                maruza = ws.cell(row=row_num, column=headers.get("Maruza (M)", 3)).value
                amaliyot = ws.cell(row=row_num, column=headers.get("Amaliyot (A)", 4)).value
                laboratoriya = ws.cell(row=row_num, column=headers.get("Laboratoriya (L)", 5)).value
                seminar = ws.cell(row=row_num, column=headers.get("Seminar (S)", 6)).value
                kurs_ishi_value = ws.cell(row=row_num, column=headers.get("Kurs ishi (K)", 7)).value
                mustaqil = ws.cell(row=row_num, column=headers.get("Mustaqil ta'lim (MT)", 8)).value
                
                # Raqamlarga o'tkazish - bo'sh bo'lsa 0
                try:
                    maruza = int(float(maruza)) if maruza else 0
                    amaliyot = int(float(amaliyot)) if amaliyot else 0
                    laboratoriya = int(float(laboratoriya)) if laboratoriya else 0
                    seminar = int(float(seminar)) if seminar else 0
                    # Kurs ishi - "Bor" bo'lsa 1, bo'sh yoki "Yo'q" bo'lsa 0
                    if kurs_ishi_value:
                        kurs_ishi_str = str(kurs_ishi_value).strip()
                        if kurs_ishi_str.lower() in ['bor', 'да', 'yes', '1', 'true']:
                            kurs_ishi = 1  # Agar mavjud bo'lsa, 1 soat deb belgilanadi
                        elif kurs_ishi_str.lower() in ["yo'q", "yoq", "нет", "no", "0", "false"]:
                            kurs_ishi = 0
                        else:
                            # Raqam formatida bo'lsa
                            kurs_ishi = int(float(kurs_ishi_str)) if kurs_ishi_str else 0
                    else:
                        # Bo'sh bo'lsa 0
                        kurs_ishi = 0
                    mustaqil = int(float(mustaqil)) if mustaqil else 0
                except (ValueError, TypeError):
                    errors.append(f"Qator {row_num}: Soatlar noto'g'ri formatda")
                    continue
                
                # Mavjud o'quv reja elementini topish yoki yaratish
                curriculum_item = DirectionCurriculum.query.filter_by(
                    direction_id=direction_id,
                    subject_id=subject.id,
                    semester=semester,
                    enrollment_year=enrollment_year,
                    education_type=education_type
                ).first()
                
                if curriculum_item:
                    # Yangilash
                    curriculum_item.hours_maruza = maruza
                    curriculum_item.hours_amaliyot = amaliyot
                    curriculum_item.hours_laboratoriya = laboratoriya
                    curriculum_item.hours_seminar = seminar
                    curriculum_item.hours_kurs_ishi = kurs_ishi
                    curriculum_item.hours_mustaqil = mustaqil
                    updated += 1
                else:
                    # Yaratish
                    curriculum_item = DirectionCurriculum(
                        direction_id=direction_id,
                        subject_id=subject.id,
                        semester=semester,
                        enrollment_year=enrollment_year,
                        education_type=education_type,
                        hours_maruza=maruza,
                        hours_amaliyot=amaliyot,
                        hours_laboratoriya=laboratoriya,
                        hours_seminar=seminar,
                        hours_kurs_ishi=kurs_ishi,
                        hours_mustaqil=mustaqil
                    )
                    db.session.add(curriculum_item)
                    imported += 1
            
            except Exception as e:
                errors.append(f"Qator {row_num}: {str(e)}")
                continue
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'updated': updated,
            'subjects_created': subjects_created,
            'errors': errors
        }
    
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'errors': [f"Fayl o'qishda xatolik: {str(e)}"]
        }


def generate_curriculum_sample_file():
    """O'quv rejani import qilish uchun namuna Excel fayl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")

    wb = Workbook()
    ws = wb.active
    ws.title = "O'quv reja"

    # Sarlavha
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "O'quv reja import uchun namuna fayl"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:I2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Eslatma va talablar
    ws['A3'] = "ESLATMA VA TALABLAR:"
    ws.merge_cells('A3:I3')
    ws['A3'].font = Font(size=11, bold=True, color="000000")
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    
    notes = [
        "1. Semestr - fan o'qitiladigan semestr (masalan: 1-semestr yoki 1)",
        "2. Fan nomi - fanning to'liq nomi (majburiy)",
        "3. Maruza (M) - maruza soatlari (butun son, mavjud bo'lsa yoziladi, yo'q bo'lsa bo'sh qoldiriladi)",
        "4. Amaliyot (A) - amaliyot soatlari (butun son, mavjud bo'lsa yoziladi, yo'q bo'lsa bo'sh qoldiriladi)",
        "5. Laboratoriya (L) - laboratoriya soatlari (butun son, mavjud bo'lsa yoziladi, yo'q bo'lsa bo'sh qoldiriladi)",
        "6. Seminar (S) - seminar soatlari (butun son, mavjud bo'lsa yoziladi, yo'q bo'lsa bo'sh qoldiriladi)",
        "7. Kurs ishi (K) - \"Bor\" yoki bo'sh (majburiy emas). Agar \"Bor\" bo'lsa, kurs ishi mavjud deb hisoblanadi. Bo'sh qoldirilsa, kurs ishi yo'q deb hisoblanadi. Kurs ishi soatlari jami soatga qo'shilmaydi.",
        "8. Mustaqil ta'lim (MT) - mustaqil ta'lim soatlari (butun son, 0 yoki bo'sh bo'lishi mumkin)",
        "9. Jami soat - avtomatik hisoblanadi (Maruza + Amaliyot + Laboratoriya + Seminar + Mustaqil ta'lim). Kurs ishi qo'shilmaydi.",
        "10. Agar fan topilmasa, yangi fan avtomatik yaratiladi.",
        "11. Bir xil semestr va fan nomi bo'lgan qatorlar yangilanadi."
    ]
    
    note_row = 4
    for note in notes:
        ws.cell(row=note_row, column=1, value=note)
        ws.merge_cells(f'A{note_row}:I{note_row}')
        ws.cell(row=note_row, column=1).font = Font(size=9, italic=True)
        ws.cell(row=note_row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        note_row += 1
    
    # Jadval sarlavhalari
    headers = [
        "Semestr",          # A
        "Fan nomi",         # B
        "Maruza (M)",       # C
        "Amaliyot (A)",     # D
        "Laboratoriya (L)", # E
        "Seminar (S)",      # F
        "Kurs ishi (K)",    # G
        "Mustaqil ta'lim (MT)", # H
        "Jami soat"         # I
    ]
    header_row = note_row + 1
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Namuna ma'lumotlar (0 soatlar bo'sh qoldiriladi, kurs ishi yo'q bo'lsa bo'sh)
    sample_data = [
        ["1-semestr", "Amaliy matematika", 60, 12, None, 15, None, 60, 147],
        ["1-semestr", "Iqtisodiy ta'limotlar tarixi", 90, None, None, 11, None, 90, 191],
        ["2-semestr", "Iqtisodiyot nazariyasi", 60, 10, None, 11, "Bor", 60, 141],
        ["4-semestr", "O'zbek (rus) tili", 60, None, None, 10, None, 60, 130],
        ["5-semestr", "O'zbek (rus) tili", 60, None, None, 10, None, 60, 130],
        ["6-semestr", "Xorijiy til", 60, None, None, 10, None, 60, 130],
        ["9-semestr", "Iqtisodiyotda axborot kommunikasiya texnol", 60, 11, 5, None, None, 60, 136]
    ]
    
    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            # None qiymatlar bo'sh qoldiriladi
            if value is not None:
                cell.value = value
            cell.alignment = Alignment(horizontal='left' if col_num <= 2 else 'center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [15, 40, 12, 12, 15, 12, 12, 18, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Jadvalni import qilish uchun namuna Excel fayl
# Jadvalni import qilish uchun namuna Excel fayl
def generate_schedule_sample_file():
    """Jadvalni import qilish uchun namuna Excel fayl (Student import kabi)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Jadval import"

    # Ustunlar
    headers = [
        'Fakultet',             # A
        'Kurs',                 # B
        'Semestr',              # C
        "Yo'nalish",            # D
        'Guruh',                # E
        'Fan',                  # F
        "O'qituvchi",           # G
        'Sana',                 # H
        'Vaqt',                 # I
        'Link',                 # J
    ]
    num_cols = len(headers)
    last_col_letter = get_column_letter(num_cols)

    # 1. Sarlavha (Title)
    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = "Dars Jadvallarini Import uchun namuna fayl"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # 2. Import talablari (Requirements)
    ws['A2'] = "IMPORT TALABLARI:"
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'].font = Font(size=11, bold=True, color="000000")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A2'].fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")

    reqs = [
        "1. Fakultet - ixtiyoriy maydon (ma'lumot uchun)",
        "2. Kurs - ixtiyoriy maydon (masalan: 1-kurs, 2-kurs)",
        "3. Semestr - ixtiyoriy maydon (masalan: 1-semestr)",
        "4. Yo'nalish - ixtiyoriy maydon (ma'lumot uchun)",
        "5. Guruh - majburiy maydon (tizimda mavjud bo'lishi shart)",
        "6. Fan - majburiy maydon (tizimda mavjud bo'lishi shart)",
        "7. O'qituvchi - majburiy maydon (Login, Passport yoki Ism-familiya)",
        "8. Sana - majburiy maydon (DD.MM.YYYY formatida)",
        "9. Vaqt - majburiy maydon (HH:MM formatida)",
        "10. Link - ixtiyoriy maydon (Zoom/Teams havolasi)"
    ]

    current_row = 3
    for req in reqs:
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = req
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        current_row += 1

    # 3. Eslatma (Notes)
    # Bo'sh qator tashlash shart emas, student importda darhol keladi, lekin o'rtada ajratuvchi yo'q. 
    # Student importda eslatma requirementsdan keyin 3 qator tashlab emas, davomidan kelmoqda (koddagi logicga kora).
    # Biz ham davom ettiramiz.
    
    ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
    note_title_cell = ws.cell(row=current_row, column=1)
    note_title_cell.value = "ESLATMA:"
    note_title_cell.font = Font(size=11, bold=True, color="000000")
    note_title_cell.alignment = Alignment(horizontal='left', vertical='center')
    note_title_cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    current_row += 1

    notes = [
        "• Fayl .xlsx formatida bo'lishi kerak",
        "• Majburiy maydonlar: Guruh, Fan, O'qituvchi, Sana, Vaqt",
        "• Guruh nomi va Fan nomi tizimdagidek aniq yozilishi kerak",
        "• Sana DD.MM.YYYY formatida (masalan: 13.01.2025)",
        "• Vaqt HH:MM formatida (masalan: 09:00)",
        "• Agar qatorda xatolik bo'lsa, o'sha qator tashlab ketiladi"
    ]

    for note in notes:
        ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = note
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
        current_row += 1

    # 4. Jadval sarlavhalari
    header_row = current_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 5. Namuna ma'lumotlar
    sample_data = [
        ['Axborot texnologiyalari', "3-kurs", "5-semestr", 'Dasturiy injiniring', 'KI-21-01', 'Oliy matematika', 'Aziz Karimov', '13.01.2025', '09:00', ''],
        ['Axborot texnologiyalari', "3-kurs", "5-semestr", 'Dasturiy injiniring', 'KI-21-01', 'Dasturlash asoslari', 'Bobur Aliyev', '13.01.2025', '11:00', 'https://zoom.us/j/123456'],
    ]

    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='center', horizontal='left')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Sana (H=8) va Vaqt (I=9) ni markazlash
            if col_num in [8, 9]:
                cell.alignment = Alignment(vertical='center', horizontal='center')

    # Ustun kengliklari
    column_widths = [20, 10, 10, 20, 15, 20, 20, 15, 10, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Excel fayldan jadvalni import qilish
def import_schedule_from_excel(file):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        errors = []
        
        # Sarlavha qatorini topish (Header Row)
        header_row_index = None
        for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
            # Qatordagi barcha qiymatlarni stringga o'tkazib tekshiramiz
            row_values = [str(cell.value).strip() if cell.value else "" for cell in row]
            if "Guruh" in row_values and "Fan" in row_values:
                header_row_index = row[0].row # 1-based index
                break
        
        if not header_row_index:
            return 0, ["Sarlavha qatori topilmadi (Guruh va Fan ustunlari bo'lishi shart)"]

        # Ma'lumotlarni o'qish
        # Headerdan keyingi qatordan boshlaymiz
        
        from app.models import Group, Subject, User, Schedule
        from sqlalchemy import func
        from datetime import datetime, time, timedelta
        
        # Ustun indekslarini aniqlash (Header row boyicha)
        header_map = {}
        for cell in ws[header_row_index]:
            if cell.value:
                header_map[str(cell.value).strip()] = cell.column - 1 # 0-based index
        
        # Kerakli ustunlar indeksi
        # Agar aniq topilmasa, default indekslarni ishlatamiz (A=0, E=4, F=5...)
        # Lekin bizning namunada: A=Fakultet, B=Kurs, C=Semestr, D=Yonalish, E=Guruh(4), F=Fan(5), G=Oqituvchi(6), H=Sana(7), I=Vaqt(8), J=Link(9)
        
        idx_group = header_map.get('Guruh', 4)
        idx_subject = header_map.get('Fan', 5)
        # O'qituvchi har xil yozilishi mumkin
        idx_teacher = header_map.get("O'qituvchi", 6)
        if "O'qituvchi" not in header_map:
             for k in header_map:
                 if "qituvchi" in k:
                     idx_teacher = header_map[k]
                     break
        
        idx_date = header_map.get('Sana (dd.mm.yyyy)', 7)
        if 'Sana (dd.mm.yyyy)' not in header_map:
             idx_date = header_map.get('Sana', 7)

        idx_time = header_map.get('Vaqt', 8)
        idx_link = header_map.get('Link', 9)

        # Iteratsiya
        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=False):
            try:
                # Helper to safe get value
                def get_val(idx):
                    if idx < len(row):
                        return row[idx].value
                    return None

                group_name = get_val(idx_group)
                subject_name = get_val(idx_subject)
                teacher_identifier = get_val(idx_teacher)
                date_val = get_val(idx_date)
                start_time_val = get_val(idx_time)
                link_val = get_val(idx_link)
                
                # Bo'sh qatorlarni o'tkazib yuborish
                if not all([group_name, subject_name, teacher_identifier, date_val, start_time_val]):
                    continue
                    
                # Guruhni topish
                group = Group.query.filter_by(name=str(group_name)).first()
                if not group:
                    errors.append(f"Qator {row[0].row}: Guruh topilmadi - {group_name}")
                    continue
                    
                # Fanni topish
                subject = Subject.query.filter_by(name=str(subject_name)).first()
                if not subject:
                    errors.append(f"Qator {row[0].row}: Fan topilmadi - {subject_name}")
                    continue
                    
                # O'qituvchini topish: Login, Passport yoki To'liq ism bo'yicha
                teacher_val = str(teacher_identifier).strip()
                teacher = None
                
                # 1. Login bo'yicha
                teacher = User.query.filter_by(login=teacher_val).first()
                # 2. Passport bo'yicha
                if not teacher:
                     teacher = User.query.filter_by(passport_number=teacher_val).first()
                # 3. To'liq ism bo'yicha (Case insensitive)
                if not teacher:
                     teacher = User.query.filter(func.lower(User.full_name) == func.lower(teacher_val)).first()
                
                if not teacher:
                    errors.append(f"Qator {row[0].row}: O'qituvchi topilmadi - {teacher_val}")
                    continue
                
                # Sana formatlash
                day_of_week_int = None
                if isinstance(date_val, datetime):
                    day_of_week_int = int(date_val.strftime('%Y%m%d'))
                elif isinstance(date_val, str):
                    for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                        try:
                            dt = datetime.strptime(date_val, fmt)
                            day_of_week_int = int(dt.strftime('%Y%m%d'))
                            break
                        except ValueError:
                            continue
                
                if not day_of_week_int:
                     errors.append(f"Qator {row[0].row}: Sana noto'g'ri formatda - {date_val}")
                     continue

                # Vaqt formatlash
                def format_time(t):
                    if isinstance(t, (datetime, time)):
                        return t.strftime('%H:%M')
                    return str(t)

                start_time = format_time(start_time_val)
                # Tugash vaqtini hisoblash (+80 min)
                end_time = ''
                try:
                    st = datetime.strptime(start_time, '%H:%M')
                    et = st + timedelta(minutes=80)
                    end_time = et.strftime('%H:%M')
                except:
                    pass
                
                # O'qituvchiga biriktirilgan barcha dars turlarini topish
                from app.models import TeacherSubject
                assigned_types = TeacherSubject.query.filter_by(
                    group_id=group.id,
                    subject_id=subject.id,
                    teacher_id=teacher.id
                ).all()
                
                types_map = {
                    'maruza': 'Ma\'ruza',
                    'lecture': 'Ma\'ruza',
                    'amaliyot': 'Amaliyot',
                    'practice': 'Amaliyot',
                    'lab': 'Laboratoriya',
                    'seminar': 'Seminar'
                }
                found_types = sorted(list(set([types_map.get(a.lesson_type, str(a.lesson_type).capitalize()) for a in assigned_types if a.lesson_type])))
                lesson_type_code = "/".join(found_types) if found_types else 'Ma\'ruza'

                # Jadval yaratish
                schedule = Schedule(
                    group_id=group.id,
                    subject_id=subject.id,
                    teacher_id=teacher.id,
                    day_of_week=day_of_week_int, # YYYYMMDD
                    start_time=start_time,
                    end_time=end_time,
                    lesson_type=lesson_type_code[:20],
                    link=link_val
                )
                db.session.add(schedule)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Qator {row[0].row}: Xatolik - {str(e)}")
                
        db.session.commit()
        return {
            'success': True,
            'imported': success_count,
            'errors': errors
        }
        
    except Exception as e:
        return {
            'success': False,
            'imported': 0,
            'errors': [f"Fayl formati noto'g'ri yoki o'qishda xatolik: {str(e)}"]
        }
