from datetime import datetime
from flask import Response
import io


def create_students_excel(students, faculty_name=None):
    """Talabalar ro'yxatini Excel formatida yaratish (yangi tartib)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"
    
    # Sarlavha
    title = f"Talabalar ro'yxati"
    if faculty_name:
        title += f" - {faculty_name}"
    
    # A–P (16 ustun)
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:P2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = [
        "Talaba ID",           # A
        "To'liq ism",          # B
        "Pasport seriya raqami",  # C
        "JSHSHIR",             # D
        "Tug'ilgan sana",      # E
        "Telefon",             # F
        "Email",               # G
        "Tavsif",              # H
        "Fakultet",            # I
        "Kurs",                # J
        "Semestr",             # K
        "Ta'lim shakli",       # L
        "Qabul yili",          # M
        "Mutaxassislik kodi",  # N
        "Mutaxassislik nomi",  # O
        "Guruh"                # P
    ]
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ma'lumotlar
    for row_num, student in enumerate(students, start=header_row + 1):
        # Talaba ID
        ws.cell(row=row_num, column=1, value=student.student_id or '')
        # To'liq ism (katta harflarda)
        ws.cell(row=row_num, column=2, value=student.full_name.upper() if student.full_name else '')
        # Pasport seriya raqami
        passport = getattr(student, 'passport_number', None)
        ws.cell(row=row_num, column=3, value=passport or '')
        # JSHSHIR
        pinfl = getattr(student, 'pinfl', None)
        ws.cell(row=row_num, column=4, value=pinfl or '')
        # Tug'ilgan sana (YYYY-MM-DD formatida)
        birth_date = getattr(student, 'birth_date', None)
        if birth_date:
            if isinstance(birth_date, str):
                ws.cell(row=row_num, column=5, value=birth_date)
            else:
                ws.cell(row=row_num, column=5, value=birth_date.strftime('%Y-%m-%d'))
        else:
            ws.cell(row=row_num, column=5, value='')
        # Telefon
        ws.cell(row=row_num, column=6, value=student.phone or '')
        # Email
        ws.cell(row=row_num, column=7, value=student.email or '')
        # Tavsif
        ws.cell(row=row_num, column=8, value=getattr(student, 'description', None) or '')
        
        # Fakultet, Kurs, Semestr, Ta'lim shakli, Guruh - guruhdan olinadi
        if getattr(student, 'group', None) and student.group:
            # Fakultet
            faculty_name_val = student.group.faculty.name if student.group.faculty else ''
            ws.cell(row=row_num, column=9, value=faculty_name_val)
            # Kurs (1-kurs formatida)
            course_year = student.group.course_year
            ws.cell(row=row_num, column=10, value=f"{course_year}-kurs" if course_year else '')
            # Semestr (talabadan yoki guruhdan) - "1-semestr" formatida
            semester = getattr(student, 'semester', None)
            if not semester and student.group:
                semester = student.group.semester
            semester_display = f"{semester}-semestr" if semester else ''
            ws.cell(row=row_num, column=11, value=semester_display)
            # Ta'lim shakli - bosh harf katta bilan
            education_type = student.group.education_type or ''
            education_type_display = education_type.capitalize() if education_type else ''
            ws.cell(row=row_num, column=12, value=education_type_display)
            # Qabul yili (yo'nalishdan yoki talabadan)
            enrollment_year = ''
            if student.group and student.group.enrollment_year:
                enrollment_year = student.group.enrollment_year
            elif getattr(student, 'enrollment_year', None):
                enrollment_year = student.enrollment_year
            ws.cell(row=row_num, column=13, value=enrollment_year)
            # Mutaxassislik kodi (yo'nalish kodi)
            specialty_code = ''
            if student.group.direction:
                specialty_code = student.group.direction.code or ''
            ws.cell(row=row_num, column=14, value=specialty_code)
            # Mutaxassislik nomi (yo'nalish nomi)
            specialty_name = ''
            if student.group.direction:
                specialty_name = student.group.direction.name or ''
            ws.cell(row=row_num, column=15, value=specialty_name)
            # Guruh
            ws.cell(row=row_num, column=16, value=student.group.name)
        else:
            # Guruh bo'lmagan talabalar uchun bo'sh qatorlar
            ws.cell(row=row_num, column=9, value='')
            ws.cell(row=row_num, column=10, value='')
            ws.cell(row=row_num, column=11, value='')
            ws.cell(row=row_num, column=12, value='')
            ws.cell(row=row_num, column=13, value='')
            ws.cell(row=row_num, column=14, value='')
            ws.cell(row=row_num, column=15, value='')
            ws.cell(row=row_num, column=16, value='')
        
        # Stil
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [15, 30, 20, 18, 18, 16, 25, 40, 20, 12, 12, 15, 12, 20, 30, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_schedule_excel(schedules, group_name=None, faculty_name=None):
    """Dars jadvalini Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dars jadvali"
    
    # Sarlavha
    title = "Dars jadvali"
    # 10 ta ustun uchun sarlavha birlashtirish (A-J)
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:J2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Agar darslar bo'lmasa
    if not schedules:
        ws['A3'] = "Darslar yo'q"
        ws.merge_cells('A3:J3')
        ws['A3'].font = Font(size=12, italic=True, color="666666")
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        # Ustun kengliklarini sozlash
        column_widths = [20, 10, 10, 20, 15, 20, 20, 15, 10, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        # Excel faylni qaytarish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # Jadval sarlavhalari
    # 1. Fakultet, 2. Kurs, 3. Semestr, 4. Yo'nalish, 5. Guruh, 6. Fan, 7. O'qituvchi, 8. Sana, 9. Vaqt, 10. Link
    # Jadval sarlavhalari - "Yangi dars qo'shish" formasi bilan bir xil
    # Asosiy: Guruh, Fan, O'qituvchi, Sana, Boshlanish vaqti, Tugash vaqti, Turi, Link
    # Info: Fakultet, Kurs, Semestr, Yo'nalish
    
    # Jadval sarlavhalari - "Yangi dars qo'shish" formasi tartibida
    # 1. Fakultet, 2. Kurs, 3. Semestr, 4. Yo'nalish, 5. Guruh, 6. Fan, 7. O'qituvchi, 8. Sana, 9. Boshlanish vaqti, 10. Link
    # (Tugash vaqti va Turi olib tashlandi)
    
    headers = [
        'Fakultet',                 # A
        'Kurs',                     # B
        'Semestr',                  # C
        "Yo'nalish",                # D
        'Guruh',                    # E
        'Fan',                      # F
        "O'qituvchi",               # G
        'Sana (dd.mm.yyyy)',        # H
        'Vaqt',                     # I
        'Link'                      # J
    ]
    header_row = 3
    
    # Sarlavhalar
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Ma'lumotlar
    for row_num, schedule in enumerate(schedules, start=header_row + 1):
        # 1. Fakultet
        faculty_val = ''
        if schedule.group and schedule.group.faculty:
            faculty_val = schedule.group.faculty.name
        ws.cell(row=row_num, column=1, value=faculty_val)
        
        # 2. Kurs
        course_val = schedule.group.course_year if schedule.group else ''
        ws.cell(row=row_num, column=2, value=f"{course_val}-kurs" if course_val else '')
        
        # 3. Semestr
        semester_val = ''
        if schedule.group:
            semester_val = schedule.group.semester
        ws.cell(row=row_num, column=3, value=f"{semester_val}-semestr" if semester_val else '')
        
        # 4. Yo'nalish
        direction_val = ''
        if schedule.group and schedule.group.direction:
            direction_val = schedule.group.direction.name
        ws.cell(row=row_num, column=4, value=direction_val)
        
        # 5. Guruh
        ws.cell(row=row_num, column=5, value=schedule.group.name if schedule.group else '')
        
        # 6. Fan
        ws.cell(row=row_num, column=6, value=schedule.subject.name if schedule.subject else '')
        
        # 7. O'qituvchi
        teacher_val = ''
        if schedule.teacher:
            # User talabiga ko'ra o'qituvchi ismi chiqariladi
            teacher_val = schedule.teacher.full_name or schedule.teacher.username
        ws.cell(row=row_num, column=7, value=teacher_val or '')
        
        # 8. Sana
        date_val = ''
        if schedule.day_of_week:
            s_date = str(schedule.day_of_week)
            if len(s_date) == 8:
                date_val = f"{s_date[6:8]}.{s_date[4:6]}.{s_date[0:4]}"
            else:
                date_val = s_date # Fallback
        ws.cell(row=row_num, column=8, value=date_val)
        
        # 9. Vaqt
        ws.cell(row=row_num, column=9, value=schedule.start_time or '')
        
        # 10. Link
        ws.cell(row=row_num, column=10, value=schedule.link or '')

        # Stil (barcha ustunlar uchun)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklari
    column_widths = [20, 10, 10, 20, 15, 25, 20, 15, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_contracts_excel(payments, course_year=None):
    """Kontrakt ma'lumotlarini Excel formatida yaratish (kurs bo'yicha)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    
    # Kurs bo'yicha guruhlash
    from collections import defaultdict
    payments_by_course = defaultdict(list)
    
    for payment in payments:
        if payment.student and payment.student.group:
            course = payment.student.group.course_year
            payments_by_course[course].append(payment)
    
    # Har bir kurs uchun alohida worksheet
    for course in sorted(payments_by_course.keys()):
        ws = wb.create_sheet(title=f"{course}-kurs")
        
        # Sarlavha
        title = f"{course}-kurs talabalar kontrakt ma'lumotlari"
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Sana
        ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Jadval sarlavhalari
        headers = ['№', 'Talaba ID', 'To\'liq ism', 'Guruh', 'Kontrakt miqdori', 'To\'lagan', 'Qolgan', 'Foiz']
        header_row = 3
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ma'lumotlar
        course_payments = payments_by_course[course]
        total_contract = 0
        total_paid = 0
        
        for row_num, payment in enumerate(course_payments, start=header_row + 1):
            student = payment.student
            contract = float(payment.contract_amount)
            paid = float(payment.paid_amount)
            remaining = contract - paid
            percentage = payment.get_payment_percentage()
            
            total_contract += contract
            total_paid += paid
            
            ws.cell(row=row_num, column=1, value=row_num - header_row)
            ws.cell(row=row_num, column=2, value=student.student_id or '')
            ws.cell(row=row_num, column=3, value=student.full_name.upper() if student.full_name else '')
            ws.cell(row=row_num, column=4, value=student.group.name if student.group else '')
            ws.cell(row=row_num, column=5, value=contract)
            ws.cell(row=row_num, column=6, value=paid)
            ws.cell(row=row_num, column=7, value=remaining)
            ws.cell(row=row_num, column=8, value=f"{percentage}%")
            
            # Stil
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Foiz bo'yicha rang
                if col_num == 8:
                    if percentage == 100:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif percentage >= 75:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
                    elif percentage >= 50:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
        
        # Jami qator
        summary_row = header_row + len(course_payments) + 2
        ws.cell(row=summary_row, column=3, value="JAMI:")
        ws.cell(row=summary_row, column=3).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=5, value=total_contract)
        ws.cell(row=summary_row, column=5).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=6, value=total_paid)
        ws.cell(row=summary_row, column=6).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=7, value=total_contract - total_paid)
        ws.cell(row=summary_row, column=7).font = Font(bold=True, size=12)
        
        # Ustun kengliklarini sozlash
        column_widths = [5, 15, 30, 15, 18, 18, 18, 10]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Bosh worksheet'ni o'chirish
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_group_grades_excel(subject, group, student_rows):
    """Guruh bo'yicha baholarni Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Baholar"
    
    # Sarlavha
    title = f"{subject.name} - {group.name} guruh baholari"
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['№', 'Talaba ID', "To'liq ism", 'Guruh', 'Fan', 'Umumiy ball', 'Maks ball', 'Foiz', 'Baho']
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    for row_num, row in enumerate(student_rows, start=header_row + 1):
        student = row['student']
        percent = row['percent']
        grade = row['grade']
        
        ws.cell(row=row_num, column=1, value=row_num - header_row)
        ws.cell(row=row_num, column=2, value=student.student_id or '')
        ws.cell(row=row_num, column=3, value=student.full_name.upper() if student.full_name else '')
        ws.cell(row=row_num, column=4, value=group.name)
        ws.cell(row=row_num, column=5, value=subject.name)
        ws.cell(row=row_num, column=6, value=row['total'])
        ws.cell(row=row_num, column=7, value=row['max_total'])
        ws.cell(row=row_num, column=8, value=f"{percent}%")
        ws.cell(row=row_num, column=9, value=f"{grade.letter} - {grade.name}" if grade else "Baholanmagan")
        
        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Baho ranglari (4-bosqichli tizim)
        percent_cell = ws.cell(row=row_num, column=8)
        if percent >= 90:
            # A - A'lo (Yashil)
            percent_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            percent_cell.font = Font(bold=True, color="006100")
        elif percent >= 70:
            # B - Yaxshi (Ko'k)
            percent_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            percent_cell.font = Font(bold=True, color="0000FF")
        elif percent >= 60:
            # C - Qoniqarli (Sariq)
            percent_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            percent_cell.font = Font(bold=True, color="9C6500")
        else:
            # D - O'tmadi (Pushti)
            percent_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            percent_cell.font = Font(bold=True, color="9C0006")
    
    # Ustun kengliklari
    widths = [5, 12, 30, 14, 24, 12, 12, 10, 16]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_all_users_excel(users):
    """Barcha foydalanuvchilarni Excel formatida yaratish (rol bo'yicha guruhlash)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    
    # Rol bo'yicha guruhlash
    from collections import defaultdict
    users_by_role = defaultdict(list)
    
    for user in users:
        # Bir nechta rol bo'lsa, har bir rol uchun alohida qo'shish
        roles = user.get_roles() if hasattr(user, 'get_roles') else [user.role]
        for role in roles:
            users_by_role[role].append(user)
    
    # Har bir rol uchun alohida worksheet
    role_names = {
        'admin': 'Administratorlar',
        'dean': 'Dekanlar',
        'teacher': "O'qituvchilar",
        'student': 'Talabalar',
        'accounting': 'Buxgalteriya'
    }
    
    for role in ['admin', 'dean', 'teacher', 'student', 'accounting']:
        if role not in users_by_role:
            continue
        
        ws = wb.create_sheet(title=role_names.get(role, role))
        
        # Sarlavha
        title = f"{role_names.get(role, role)} ro'yxati"
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Sana
        ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:K2')
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Jadval sarlavhalari
        if role == 'student':
            headers = ['№', "To'liq ism", 'Email', 'Telefon', 'Talaba ID', 'Pasport raqami', 'JSHSHIR', 'Tug\'ilgan sana', 'Guruh', 'Kurs', 'Fakultet']
        else:
            headers = ['№', "To'liq ism", 'Email', 'Telefon', 'Pasport raqami', 'JSHSHIR', 'Tug\'ilgan sana', 'Kafedra', 'Lavozim', 'Fakultet', 'Holat']
        
        header_row = 3
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Ma'lumotlar
        role_users = users_by_role[role]
        for row_num, user in enumerate(role_users, start=header_row + 1):
            ws.cell(row=row_num, column=1, value=row_num - header_row)
            ws.cell(row=row_num, column=2, value=user.full_name.upper() if user.full_name else '')
            ws.cell(row=row_num, column=3, value=user.email)
            ws.cell(row=row_num, column=4, value=user.phone or '')
            
            if role == 'student':
                ws.cell(row=row_num, column=5, value=user.student_id or '')
                ws.cell(row=row_num, column=6, value=getattr(user, 'passport_number', None) or '')
                ws.cell(row=row_num, column=7, value=getattr(user, 'pinfl', None) or '')
                birth_date = getattr(user, 'birth_date', None)
                ws.cell(row=row_num, column=8, value=birth_date.strftime('%Y-%m-%d') if birth_date else '')
                ws.cell(row=row_num, column=9, value=user.group.name if user.group else '')
                ws.cell(row=row_num, column=10, value=user.group.course_year if user.group else '')
                ws.cell(row=row_num, column=11, value=user.group.faculty.name if user.group and user.group.faculty else '')
            else:
                ws.cell(row=row_num, column=5, value=getattr(user, 'passport_number', None) or '')
                ws.cell(row=row_num, column=6, value=getattr(user, 'pinfl', None) or '')
                birth_date = getattr(user, 'birth_date', None)
                ws.cell(row=row_num, column=7, value=birth_date.strftime('%Y-%m-%d') if birth_date else '')
                ws.cell(row=row_num, column=8, value=user.department or '')
                ws.cell(row=row_num, column=9, value=user.position or '')
                ws.cell(row=row_num, column=10, value=user.managed_faculty.name if user.managed_faculty else '')
                ws.cell(row=row_num, column=11, value='Faol' if user.is_active else 'Bloklangan')
            
            # Stil
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ustun kengliklarini sozlash
        if role == 'student':
            column_widths = [5, 30, 25, 16, 15, 18, 16, 14, 14, 8, 20]
        else:
            column_widths = [5, 30, 25, 16, 18, 16, 14, 20, 15, 20, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Bosh worksheet'ni o'chirish
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_staff_excel(users):
    """Xodimlarni Excel formatida yaratish (bitta sheet'da) - bir nechta rollarni qo'llab-quvvatlash"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Xodimlar"
    
    # Sarlavha
    title = "Xodimlar ro'yxati"
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:I2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari (A ustunidan boshlanadi)
    headers = ["To'liq ism", 'Login', 'Pasport seriya raqami', 'JSHSHIR', "Tug'ilgan sana", 'Telefon', 'Email', 'Tavsif', 'Rollar']
    
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ma'lumotlar
    for row_num, user in enumerate(users, start=header_row + 1):
        ws.cell(row=row_num, column=1, value=user.full_name.upper() if user.full_name else '')
        ws.cell(row=row_num, column=2, value=user.login or '')
        ws.cell(row=row_num, column=3, value=getattr(user, 'passport_number', None) or '')
        ws.cell(row=row_num, column=4, value=getattr(user, 'pinfl', None) or '')
        
        # Tug'ilgan sana (YYYY-MM-DD formatida)
        birth_date = getattr(user, 'birth_date', None)
        if birth_date:
            if isinstance(birth_date, str):
                ws.cell(row=row_num, column=5, value=birth_date)
            else:
                ws.cell(row=row_num, column=5, value=birth_date.strftime('%Y-%m-%d'))
        else:
            ws.cell(row=row_num, column=5, value='')
        
        ws.cell(row=row_num, column=6, value=user.phone or '')
        ws.cell(row=row_num, column=7, value=user.email or '')
        ws.cell(row=row_num, column=8, value=getattr(user, 'description', None) or '')
        
        # Rollarni olish va o'zbek tilida ko'rsatish (belgilangan tartibda)
        from app.models import UserRole
        role_names = {
            'admin': 'Administrator',
            'dean': 'Dekan',
            'teacher': "O'qituvchi",
            'accounting': 'Buxgalter',
            'student': 'Talaba'
        }
        # Belgilangan tartib: Administrator, Dekan, O'qituvchi, Buxgalter
        role_order = ['admin', 'dean', 'teacher', 'accounting', 'student']
        role_display_order = ['Administrator', 'Dekan', "O'qituvchi", 'Buxgalter', 'Talaba']
        
        # Foydalanuvchining rollarini olish
        user_roles = UserRole.query.filter_by(user_id=user.id).all()
        user_role_codes = [ur.role for ur in user_roles] if user_roles else ([user.role] if user.role else [])
        
        # Belgilangan tartibda ko'rsatish
        roles_display = []
        for role_code in role_order:
            if role_code in user_role_codes:
                roles_display.append(role_names.get(role_code, role_code))
        
        roles_str = ', '.join(roles_display) if roles_display else ''
        ws.cell(row=row_num, column=9, value=roles_str)
        
        # Stil
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [30, 20, 20, 18, 18, 16, 25, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


def create_sample_contracts_excel():
    """Kontrakt import uchun namuna Excel fayl yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Namuna"
    
    # Sarlavha
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "Kontrakt ma'lumotlarini import qilish uchun namuna"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sarlavha qatori
    headers = ['Talaba_id', 'Ismi', 'Kontrakt miqdori', 'To\'lagani']
    header_row = 2
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Namuna ma'lumotlar
    sample_data = [
        ['376100000000000', 'Tursunov Avazbek', '12652200', '6520020'],
        ['376100000000001', 'Karimova Malika', '12652200', '12652200'],
        ['376100000000002', 'Rahimov Dilshod', '12652200', '6326100'],
        ['376100000000003', 'Aliyeva Nodira', '12652200', '0'],
    ]
    
    for row_num, row_data in enumerate(sample_data, start=header_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Eslatma qatori
    note_row = header_row + len(sample_data) + 2
    ws.merge_cells(f'A{note_row}:D{note_row}')
    note_cell = ws[f'A{note_row}']
    note_cell.value = "ESLATMA: Talaba_id yoki Ismi orqali talaba topiladi. Kontrakt miqdori majburiy."
    note_cell.font = Font(size=10, italic=True, color="666666")
    note_cell.alignment = Alignment(horizontal='left', vertical='center')
    note_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [20, 30, 18, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_subjects_excel(subjects):
    """Fanlar ro'yxatini Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Fanlar"
    
    # Sarlavha
    title = "Fanlar ro'yxati"
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:B2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari
    headers = [
        "Fan nomi",      # A
        "Tavsif"         # B
    ]
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ma'lumotlar
    for row_num, subject in enumerate(subjects, start=header_row + 1):
        ws.cell(row=row_num, column=1, value=subject.name)
        ws.cell(row=row_num, column=2, value=subject.description or '')
        
        # Stil
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ustun kengliklarini sozlash
    column_widths = [40, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_curriculum_excel(direction, curriculum_items):
    """O'quv rejani Excel formatida yaratish (rasmdagi formatga mos)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. Iltimos, 'pip install openpyxl' buyrug'ini bajaring.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "O'quv reja"
    
    # Sarlavha
    title = f"O'quv reja - {direction.code} - {direction.name}"
    # Enrollment year direction o'rniga guruhdan olinadi, lekin bu yerda export funksiyasi direction ob'ektini oladi
    # Bu yerda direction nomi saqlanib qoladi. Enrollment year direction modelida yo'q.
    pass
    
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A2:I2')
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Jadval sarlavhalari (rasmdagi tartibda)
    headers = [
        "Semestr",          # A
        "Fan nomi",         # B
        "Maruza (M)",       # C
        "Amaliyot (A)",     # D
        "Laboratoriya (L)", # E
        "Seminar (S)",      # F
        "Kurs ishi (K)",    # G
        "Mustaqil ta'lim (MT)", # H
        "Jami soat"         # I
    ]
    header_row = 3
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ma'lumotlarni semestr bo'yicha tartiblash
    sorted_items = sorted(curriculum_items, key=lambda x: (x.semester, x.subject.name))
    
    total_maruza = 0
    total_amaliyot = 0
    total_laboratoriya = 0
    total_seminar = 0
    total_kurs_ishi = 0
    total_mustaqil = 0
    total_jami = 0
    
    for row_num, item in enumerate(sorted_items, start=header_row + 1):
        subject = item.subject
        maruza = item.hours_maruza or 0
        amaliyot = item.hours_amaliyot or 0
        laboratoriya = item.hours_laboratoriya or 0
        seminar = item.hours_seminar or 0
        kurs_ishi = item.hours_kurs_ishi or 0
        mustaqil = item.hours_mustaqil or 0
        # Kurs ishi jami soatga qo'shilmaydi
        jami = maruza + amaliyot + laboratoriya + seminar + mustaqil
        
        total_maruza += maruza
        total_amaliyot += amaliyot
        total_laboratoriya += laboratoriya
        total_seminar += seminar
        total_kurs_ishi += kurs_ishi
        total_mustaqil += mustaqil
        total_jami += jami
        
        # Semestr (1-semestr formatida)
        ws.cell(row=row_num, column=1, value=f"{item.semester}-semestr")
        # Fan nomi
        ws.cell(row=row_num, column=2, value=subject.name)
        # Soatlar - 0 bo'lsa bo'sh qoldiriladi
        ws.cell(row=row_num, column=3, value=maruza if maruza > 0 else None)
        ws.cell(row=row_num, column=4, value=amaliyot if amaliyot > 0 else None)
        ws.cell(row=row_num, column=5, value=laboratoriya if laboratoriya > 0 else None)
        ws.cell(row=row_num, column=6, value=seminar if seminar > 0 else None)
        # Kurs ishi - bor bo'lsa "Bor", yo'q bo'lsa bo'sh
        kurs_ishi_text = "Bor" if kurs_ishi > 0 else None
        ws.cell(row=row_num, column=7, value=kurs_ishi_text)
        ws.cell(row=row_num, column=8, value=mustaqil if mustaqil > 0 else None)
        ws.cell(row=row_num, column=9, value=jami)
        
        # Stil
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left' if col_num <= 2 else 'center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Jami qator
    summary_row = header_row + len(sorted_items) + 2
    ws.cell(row=summary_row, column=1, value="JAMI:")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=3, value=total_maruza)
    ws.cell(row=summary_row, column=3).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=4, value=total_amaliyot)
    ws.cell(row=summary_row, column=4).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=5, value=total_laboratoriya)
    ws.cell(row=summary_row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=6, value=total_seminar)
    ws.cell(row=summary_row, column=6).font = Font(bold=True, size=12)
    # Kurs ishi jami - faqat "Bor" yoki "Yo'q" ko'rsatiladi
    ws.cell(row=summary_row, column=7, value="")
    ws.cell(row=summary_row, column=7).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=8, value=total_mustaqil)
    ws.cell(row=summary_row, column=8).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=9, value=total_jami)
    ws.cell(row=summary_row, column=9).font = Font(bold=True, size=12)
    
    # Ustun kengliklarini sozlash
    column_widths = [15, 40, 12, 12, 15, 12, 12, 18, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Excel faylni qaytarish
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_detailed_assignment_export_excel(subject, group, assignments, matrix):
    """Guruh bo'yicha batafsil topshiriq baholarini (ustunma-ustun) Excel formatida yaratish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Batafsil Baholar"
    
    # Sarlavha
    title = f"{subject.name} - {group.name} guruhining batafsil baholari"
    last_col_letter = get_column_letter(len(assignments) + 4)
    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sana
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # 1-qator sarlavhalari (Titles)
    headers1 = ['Talaba ID', "To'liq ism"]
    for a in assignments:
        headers1.append(a.title)
    headers1.append('Jami ball')
    headers1.append('O\'zlashtirish (%)')
    
    # 2-qator sarlavhalari (Sub-info)
    headers2 = ['', '']
    total_max = 0
    for a in assignments:
        lesson_type_display = {
            'maruza': 'Maruza',
            'amaliyot': 'Amaliyot',
            'laboratoriya': 'Laboratoriya',
            'seminar': 'Seminar',
            'kurs_ishi': 'Kurs ishi'
        }.get(a.lesson_type, a.lesson_type or '-')
        headers2.append(f"{lesson_type_display} ({a.max_score})")
        total_max += (a.max_score or 0)
    headers2.append(f"Maks: {total_max}")
    headers2.append("100%")
    
    header_row1 = 3
    header_row2 = 4
    
    # Stil sozlamalari
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_white = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Fill colors for grades
    fill_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
    fill_b = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Blue
    fill_c = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
    fill_d = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Pink
    
    font_a = Font(bold=True, color="006100")
    font_b = Font(bold=True, color="0000FF")
    font_c = Font(bold=True, color="9C6500")
    font_d = Font(bold=True, color="9C0006")
    
    for col_num, val in enumerate(headers1, 1):
        cell = ws.cell(row=header_row1, column=col_num, value=val)
        cell.font = font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        
    for col_num, val in enumerate(headers2, 1):
        cell = ws.cell(row=header_row2, column=col_num, value=val)
        cell.font = font_bold
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    # Ma'lumotlar
    for row_num, row_data in enumerate(matrix, start=5):
        # A: ID
        ws.cell(row=row_num, column=1, value=row_data['student_id']).border = border_thin
        # B: Name
        ws.cell(row=row_num, column=2, value=row_data['student_name']).border = border_thin
        
        # C onwards: Scores
        for col_idx, score in enumerate(row_data['scores'], start=3):
            cell = ws.cell(row=row_num, column=col_idx, value=score)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
            # Score ranglari
            if assignments[col_idx-3].max_score:
                ratio = (score / assignments[col_idx-3].max_score) * 100
                if ratio >= 90: cell.fill = fill_a
                elif ratio >= 70: cell.fill = fill_b
                elif ratio >= 60: cell.fill = fill_c
                else: cell.fill = fill_d
        
        # Next: Total
        total_col = len(assignments) + 3
        cell_total = ws.cell(row=row_num, column=total_col, value=row_data['total_score'])
        cell_total.font = font_bold
        cell_total.alignment = Alignment(horizontal='center')
        cell_total.border = border_thin
        
        # Last: Percent
        percent_col = len(assignments) + 4
        cell_percent = ws.cell(row=row_num, column=percent_col, value=f"{row_data['percent']}%")
        cell_percent.font = font_bold
        cell_percent.alignment = Alignment(horizontal='center')
        cell_percent.border = border_thin
        # Ranglar
        if row_data['percent'] >= 90:
            cell_percent.fill = fill_a
            cell_percent.font = font_a
        elif row_data['percent'] >= 70:
            cell_percent.fill = fill_b
            cell_percent.font = font_b
        elif row_data['percent'] >= 60:
            cell_percent.fill = fill_c
            cell_percent.font = font_c
        else:
            cell_percent.fill = fill_d
            cell_percent.font = font_d

    # Ustun kengliklari
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    for i in range(3, percent_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
